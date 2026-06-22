from flask import Blueprint, request, jsonify, send_file
from app import db
from app.models.tariff_library import TariffLibrary
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io

tariff_bp = Blueprint('tariff', __name__)

@tariff_bp.route('/api/tariff/template', methods=['GET'])
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tariff Template'
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    headers = ['Rate Name', 'Rate Value (PHP/kWh)', 'Unit', 'Description']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin
    example = ['Residential Rate', '12.50', 'PHP/kWh', 'Example rate']
    for i, v in enumerate(example, 1):
        ws.cell(row=2, column=i, value=v).border = thin
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='tariff_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@tariff_bp.route('/api/tariff/import', methods=['POST'])
def import_tariff():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    wb = Workbook()
    ws = wb.active
    # Load uploaded file
    wb = __import__('openpyxl').load_workbook(f)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            t = TariffLibrary(
                rate_name=str(row[0]).strip(),
                rate_value=float(row[1]),
                unit=str(row[2]).strip() if row[2] else 'PHP/kWh',
                description=str(row[3]).strip() if row[3] else ''
            )
            db.session.add(t)
            count += 1
    db.session.commit()
    return jsonify({'ok': True, 'imported': count})

@tariff_bp.route('/api/tariff', methods=['GET'])
def list_tariff():
    items = TariffLibrary.query.order_by(TariffLibrary.rate_name).all()
    return jsonify([i.to_dict() for i in items])

@tariff_bp.route('/api/tariff', methods=['POST'])
def add_tariff():
    data = request.get_json()
    t = TariffLibrary(
        rate_name=data['rate_name'],
        rate_value=float(data['rate_value']),
        unit=data.get('unit', 'PHP/kWh'),
        description=data.get('description', '')
    )
    db.session.add(t)
    db.session.commit()
    return jsonify(t.to_dict()), 201

@tariff_bp.route('/api/tariff/<int:tid>', methods=['DELETE'])
def del_tariff(tid):
    t = TariffLibrary.query.get_or_404(tid)
    db.session.delete(t)
    db.session.commit()
    return jsonify({'ok': True})

