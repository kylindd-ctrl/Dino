import logging
from flask import Blueprint, request, jsonify
from app import db
from app.models import Project
from app.services.map_service import parse_google_maps_url, resolve_maps_url

logger = logging.getLogger(__name__)
projects_bp = Blueprint("projects", __name__)

@projects_bp.route("", methods=["GET"])
def list_projects():
    projects = Project.query.order_by(Project.updated_at.desc()).all()
    return jsonify([p.to_dict() for p in projects])

@projects_bp.route("", methods=["POST"])
def create_project():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Request body required"}), 400
    name = data.get("name", "").strip()
    customer = data.get("customer_name", "").strip()
    maps_link = data.get("google_maps_link", "").strip()
    if not name:
        return jsonify({"error": "Project name is required"}), 400

    # Parse Google Maps link (now handles redirects)
    geo = parse_google_maps_url(maps_link)

    project = Project(
        name=name,
        customer_name=customer,
        google_maps_link=maps_link,
        latitude=geo["latitude"],
        longitude=geo["longitude"],
        address=geo["address"],
        status="draft",
    )
    db.session.add(project)
    db.session.commit()
    logger.info("Created project %d: %s", project.id, project.name)
    return jsonify(project.to_dict()), 201

@projects_bp.route("/<int:project_id>", methods=["GET"])
def get_project(project_id):
    project = Project.query.get_or_404(project_id)
    data = project.to_dict()
    if project.pv_system:
        data["pv_system"] = project.pv_system.to_dict()
    if project.modules:
        data["modules"] = [m.to_dict() for m in project.modules]
    if project.inverters:
        data["inverters"] = [i.to_dict() for i in project.inverters]
    if project.solar_resource:
        data["solar_resource"] = project.solar_resource.to_dict()
    if project.financial_results:
        data["financial_results"] = [r.to_dict() for r in project.financial_results]
    return jsonify(data)

@projects_bp.route("/<int:project_id>", methods=["PUT"])
def update_project(project_id):
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    if "name" in data:
        project.name = data["name"]
    if "customer_name" in data:
        project.customer_name = data["customer_name"]
    if "google_maps_link" in data:
        project.google_maps_link = data["google_maps_link"]
        geo = parse_google_maps_url(data["google_maps_link"])
        project.latitude = geo["latitude"]
        project.longitude = geo["longitude"]
        project.address = geo["address"]
    if "status" in data:
        project.status = data["status"]
    if "exchange_rate" in data:
        project.exchange_rate = float(data["exchange_rate"])
    if "electricity_rate" in data:
        project.electricity_rate = float(data["electricity_rate"])
    if "revenue_discount" in data:
        project.revenue_discount = float(data["revenue_discount"])
    if "total_revenue_php" in data:
        project.total_revenue_php = float(data["total_revenue_php"])
    if "total_investment_php" in data:
        project.total_investment_php = float(data["total_investment_php"])
    if "roof_area" in data:
        project.roof_area = float(data["roof_area"]) if data["roof_area"] else None
    if "exchg_date" in data:
        project.exchg_date = data["exchg_date"]
    db.session.commit()
    return jsonify(project.to_dict())

@projects_bp.route("/<int:project_id>", methods=["DELETE"])
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    # Cascade delete all related records (SQLite FK constraint handling)
    from app.models import PVSystem, PVModule, Inverter, SolarResource
    from app.models import MonthlyGeneration, HourlyGeneration, FinancialResult, Upload, QuoteItem

    for m in PVModule.query.filter_by(project_id=project.id).all():
        db.session.delete(m)
    for inv in Inverter.query.filter_by(project_id=project.id).all():
        db.session.delete(inv)
    for u in Upload.query.filter_by(project_id=project.id).all():
        db.session.delete(u)
    for fr in FinancialResult.query.filter_by(project_id=project.id).all():
        db.session.delete(fr)
    for mg in MonthlyGeneration.query.filter_by(project_id=project.id).all():
        db.session.delete(mg)
    for hg in HourlyGeneration.query.filter_by(project_id=project.id).all():
        db.session.delete(hg)
    if project.pv_system:
        db.session.delete(project.pv_system)
    for qi in QuoteItem.query.filter_by(project_id=project.id).all():
        db.session.delete(qi)
    if project.solar_resource:
        db.session.delete(project.solar_resource)

    db.session.delete(project)
    db.session.commit()
    return jsonify({"message": "Deleted"})

# --- New: resolve Google Maps link (handles short URLs) ---
@projects_bp.route("/resolve-link", methods=["POST"])
def resolve_link():
    """Resolve a Google Maps URL (including short links) and return coordinates."""
    data = request.get_json()
    if not data or not data.get("url"):
        return jsonify({"error": "URL required"}), 400

    result = resolve_maps_url(data["url"])
    return jsonify(result)



@projects_bp.route("/<int:pid>/fetch-rate", methods=["GET"])
def fetch_rate(pid):
    """Fetch current PHP exchange rate."""
    import urllib.request, json
    try:
        # Use exchangerate-api.com free API
        url = "https://api.exchangerate-api.com/v4/latest/CNY"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=5)
        data = json.loads(resp.read())
        if data and "rates" in data and "PHP" in data["rates"]:
            rate = data["rates"]["PHP"]
            return jsonify({"rate": round(rate, 2)})
    except Exception as e:
        pass
    return jsonify({"rate": 8.76, "note": "Using default rate"})

@projects_bp.route("/<int:pid>/export-php-quote")
def export_php_quote(pid):
    """Export PHP Quotation XLSX."""
    project = Project.query.get_or_404(pid)
    from app.models import QuoteItem
    items = QuoteItem.query.filter_by(project_id=pid).order_by(QuoteItem.id).all()
    import openpyxl, io
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from flask import send_file
    rate = project.exchange_rate or 8.76
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PHP Quotation"
    thin = Border(*[Side(style="thin")]*4)
    for ci, h in enumerate(["No","Description","Quantity","Unit Price","Unit","Total Price"], 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    total_val = 0
    for idx_i, item in enumerate(items, 1):
        r = idx_i + 1
        cny_total = item.total_price()
        php_up = round(cny_total * rate / item.quantity) if item.quantity else 0
        vals = [idx_i, item.equipment_name, item.quantity, php_up, item.unit or "pcs", round(cny_total * rate)]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = thin
            if ci == 2: cell.alignment = Alignment(wrap_text=True)
            if ci in (3,6): cell.alignment = Alignment(horizontal="center")
        total_val += round(item.total_price() * rate)
    tr = len(items) + 2
    for ci in range(1, 7): ws.cell(row=tr, column=ci).border = thin
    ws.cell(row=tr, column=2, value="Total").font = Font(bold=True)
    ws.cell(row=tr, column=6, value=total_val).font = Font(bold=True)
    for w, cw in [["A",6],["B",35],["C",10],["D",18],["E",10],["F",18]]:
        ws.column_dimensions[w].width = cw
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    nm = (project.name or "quotation").replace(" ", "_")
    return send_file(bio, as_attachment=True, download_name=nm + "_PHP_Quotation.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

