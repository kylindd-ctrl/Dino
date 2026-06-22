from flask import Blueprint, request, jsonify
from app import db
from app.models import Project, PriceLibrary, QuoteItem, MonthlyGeneration

quotation_bp = Blueprint("quotation", __name__)

# Price Library CRUD
@quotation_bp.route("/api/price-library", methods=["GET"])
def list_lib():
    items = PriceLibrary.query.order_by(PriceLibrary.equipment_name).all()
    return jsonify([i.to_dict() for i in items])

@quotation_bp.route("/api/price-library", methods=["POST"])
def add_lib():
    data = request.get_json()
    i = PriceLibrary(equipment_name=data["equipment_name"], model=data.get("model",""),
                     unit_price=float(data["unit_price"]), unit=data.get("unit","pcs"))
    db.session.add(i)
    db.session.commit()
    return jsonify(i.to_dict()), 201

@quotation_bp.route("/api/price-library/<int:lid>", methods=["DELETE"])
def del_lib(lid):
    i = PriceLibrary.query.get_or_404(lid)
    db.session.delete(i)
    db.session.commit()
    return jsonify({"ok": True})

# Quote Items CRUD
@quotation_bp.route("/api/projects/<int:pid>/quote", methods=["GET"])
def list_quote(pid):
    items = QuoteItem.query.filter_by(project_id=pid).order_by(QuoteItem.id).all()
    return jsonify([i.to_dict() for i in items])

@quotation_bp.route("/api/projects/<int:pid>/quote", methods=["POST"])
def add_quote(pid):
    data = request.get_json()
    i = QuoteItem(project_id=pid, equipment_name=data["equipment_name"],
                  model=data.get("model",""), unit_price=float(data["unit_price"]),
                  quantity=int(data.get("quantity",1)),
                  gross_margin=float(data.get("gross_margin",0)),
                  unit=data.get("unit","pcs"))
    db.session.add(i)
    db.session.commit()
    return jsonify(i.to_dict()), 201

@quotation_bp.route("/api/projects/<int:pid>/quote/<int:iid>", methods=["PUT"])
def upd_quote(pid, iid):
    i = QuoteItem.query.get_or_404(iid)
    data = request.get_json()
    for attr in ["equipment_name","model","unit_price","quantity","gross_margin","unit"]:
        if attr in data:
            setattr(i, attr, float(data[attr]) if attr in ["unit_price","gross_margin"] else int(data[attr]) if attr == "quantity" else data[attr])
    db.session.commit()
    return jsonify(i.to_dict())

@quotation_bp.route("/api/projects/<int:pid>/quote/<int:iid>", methods=["DELETE"])
def del_quote(pid, iid):
    i = QuoteItem.query.get_or_404(iid)
    db.session.delete(i)
    db.session.commit()
    return jsonify({"ok": True})

# Revenue - GET (load data)
@quotation_bp.route("/api/projects/<int:pid>/revenue", methods=["GET"])
def get_revenue(pid):
    p = Project.query.get_or_404(pid)
    monthly = MonthlyGeneration.query.filter_by(project_id=pid).order_by(MonthlyGeneration.month).all()
    days_list = [31,28,31,30,31,30,31,31,30,31,30,31]
    mdata = {m.month: m for m in monthly}
    months = []
    for month in range(1, 13):
        m = mdata.get(month)
        if m:
            pvtotal = m.pvtotal_kwh or 0
            days = days_list[month-1]
            dg = round(pvtotal / days, 2) if days else 0
            wd = m.working_days if m.working_days is not None and m.working_days > 0 else days
            rev = m.revenue_php if m.revenue_php is not None and m.revenue_php > 0 else round(wd * dg * (p.electricity_rate or 13), 2) if wd and dg else 0
        else:
            dg = 0
            wd = 0
            rev = 0
        months.append({"month": month, "daily_gen": dg, "working_days": wd, "revenue_php": rev})
    total = sum(x["revenue_php"] for x in months) * (p.revenue_discount or 1.0)
    return jsonify({"months": months, "discount": (p.revenue_discount or 1.0) * 100, "rate": p.electricity_rate or 13, "total_revenue": round(total, 2)})

# Revenue - PUT (save data) - THIS WAS THE MISSING ROUTE!
@quotation_bp.route("/api/projects/<int:pid>/revenue", methods=["PUT"])
def save_revenue(pid):
    p = Project.query.get_or_404(pid)
    data = request.get_json()
    if "discount" in data:
        p.revenue_discount = float(data["discount"]) / 100
    if "rate" in data:
        p.electricity_rate = float(data["rate"])
    if "working_days" in data:
        days_list = [31,28,31,30,31,30,31,31,30,31,30,31]
        for idx, wd in enumerate(data["working_days"], 1):
            mg = MonthlyGeneration.query.filter_by(project_id=pid, month=idx).first()
            if mg:
                mg.working_days = int(wd)
                rev = round(int(wd) * (mg.pvtotal_kwh or 0) * (p.electricity_rate or 13) / days_list[idx-1], 2) if int(wd) and mg.pvtotal_kwh else 0
                mg.revenue_php = rev
    total_rev = sum(mg.revenue_php or 0 for mg in MonthlyGeneration.query.filter_by(project_id=pid).all()) * (p.revenue_discount or 1.0)
    p.total_revenue_php = round(total_rev, 2)
    db.session.commit()
    return jsonify({"ok": True, "total_revenue": p.total_revenue_php})

# Fetch exchange rate (using exchangerate.host free API)
@quotation_bp.route("/api/projects/<int:pid>/fetch-rate", methods=["GET"])
def fetch_rate(pid):
    import requests
    try:
        resp = requests.get("https://api.exchangerate-api.com/v4/latest/CNY", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            php_rate = data["rates"].get("PHP")
            if php_rate:
                p = Project.query.get_or_404(pid)
                p.exchange_rate = round(php_rate, 4)
                db.session.commit()
                return jsonify({"rate": round(php_rate, 4)})
        return jsonify({"rate": 8.76})
    except Exception as e:
        return jsonify({"rate": 8.76, "note": str(e)})

# Export PHP Quotation
@quotation_bp.route("/api/projects/<int:pid>/export-php-quote", methods=["GET"])
def export_php_quote(pid):
    import openpyxl, io
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    p = Project.query.get_or_404(pid)
    items = QuoteItem.query.filter_by(project_id=pid).order_by(QuoteItem.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PHP Quotation"
    rate = getattr(p, 'exchange_rate', None) or 8.76
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for col, h in enumerate(["#","Description","Quantity","Unit Price","Unit","Total Price"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = thin
    for idx, i in enumerate(items, 1):
        r = idx + 1
        vals = [idx, i.equipment_name, i.quantity, round(i.unit_price * rate, 2), i.unit or "pcs", round(i.total_price() * rate, 2)]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v).border = thin
    tr = len(items) + 2
    for ci in range(1, 7): ws.cell(row=tr, column=ci).border = thin
    ws.cell(row=tr, column=2, value="Total").font = Font(bold=True)
    ws.cell(row=tr, column=6, value=round(sum(i.total_price() for i in items) * rate, 2)).font = Font(bold=True)
    for w, c in [("A",6),("B",35),("C",10),("D",18),("E",10),("F",18)]:
        ws.column_dimensions[w].width = c
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    from flask import send_file
    nm = (p.name or "quotation").replace(" ", "_")
    return send_file(bio, as_attachment=True, download_name=nm + "_PHP_Quotation.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
